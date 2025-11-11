# app.py
# ------------------------------------------------------------
# 故障メール → 正規表現抽出 → 既存テンプレ(.xlsm)へ書込み → ダウンロード
# 3ステップUI / パスコード認証 / 編集不可 / 折りたたみ表示（時系列）
# 仕様反映：
#   - 曜日：日本語（例：月）
#   - 複数行：最大5行。超過は「…」付与
#   - 通報者：原文そのまま（様/電話番号含む）
#   - ファイル名：管理番号_物件名_日付（yyyymmdd）
#   - マクロ保持対応（keep_vba=True）
# ------------------------------------------------------------
import io
import re
import unicodedata
from datetime import datetime, timedelta, timezone
from typing import Dict, Optional, Tuple, List
import os
import sys
import traceback
import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage  # 画像機能は将来用
import streamlit as st

# ---- 基本設定 ------------------------------------------------
JST = timezone(timedelta(hours=9))
APP_TITLE = "故障報告書自動生成"

def _get_passcode() -> str:
    try:
        val = st.secrets.get("APP_PASSCODE")
        if val:
            return str(val)
    except Exception:
        pass
    env_val = os.getenv("APP_PASSCODE")
    if env_val:
        return str(env_val)
    return ""  # 開発用

SHEET_NAME = "緊急出動報告書（リンク付き）"
WEEKDAYS_JA = ["月", "火", "水", "木", "金", "土", "日"]

# -------------------------------------------------------------
# ✏️ 編集フィールド共通関数
# -------------------------------------------------------------
def _ensure_extracted():
    if "extracted" not in st.session_state or st.session_state.extracted is None:
        st.session_state.extracted = {}

def _enter_edit_mode():
    _ensure_extracted()
    st.session_state.edit_mode = True
    st.session_state.edit_buffer = copy.deepcopy(st.session_state.extracted)

def _cancel_edit():
    st.session_state.edit_mode = False
    st.session_state.edit_buffer = {}

def _save_edit():
    st.session_state.extracted = copy.deepcopy(st.session_state.edit_buffer)
    st.session_state.edit_mode = False
    st.session_state.edit_buffer = {}

def _get_working_dict() -> dict:
    if st.session_state.get("edit_mode"):
        return st.session_state.edit_buffer
    return st.session_state.extracted or {}

def _set_working_value(key: str, value: str):
    if st.session_state.get("edit_mode"):
        st.session_state.edit_buffer[key] = value
    else:
        _ensure_extracted()
        st.session_state.extracted[key] = value

# ✅ 必須（編集可能項目=必須）
REQUIRED_KEYS = [
    "通報者", "受信内容", "現着状況", "原因", "処置内容", "処理修理後", "所属",
]

def _is_required_missing(data: dict, key: str) -> bool:
    return key in REQUIRED_KEYS and not (data.get(key) or "").strip()

def _display_text(value: str, max_lines: int):
    if not value:
        return ""
    if max_lines and max_lines > 1:
        lines = _split_lines(value, max_lines=max_lines)
        return "<br>".join(lines)
    return value.replace("\n", "<br>")

def render_field(label: str, key: str, max_lines: int = 1, placeholder: str = "", editable_in_bulk: bool = False):
    data = _get_working_dict()
    val = data.get(key) or ""
    missing = _is_required_missing(data, key)

    cols = st.columns([0.22, 0.78])
    with cols[0]:
        st.markdown(("🔴 **" if missing else "**") + f"{label}**")

    with cols[1]:
        if st.session_state.get("edit_mode") and editable_in_bulk:
            if max_lines == 1:
                new_val = st.text_input("", value=val, placeholder=placeholder, key=f"in_{key}")
            else:
                new_val = st.text_area("", value=val, placeholder=placeholder, height=max(80, max_lines * 24), key=f"ta_{key}")
            _set_working_value(key, new_val)
        else:
            st.markdown("<span class='missing'>未入力</span>" if missing else _display_text(val, max_lines=max_lines),
                        unsafe_allow_html=True)

# ====== テキスト整形ユーティリティ ======
def normalize_text(text: str) -> str:
    if not text:
        return ""
    t = unicodedata.normalize("NFKC", text)
    t = t.replace("：", ":")  # コロン統一（NFKCでも残る場合に備え二重化）
    t = t.replace("\t", " ").replace("\r\n", "\n").replace("\r", "\n")
    # 余計な全角空白を半角へ（NFKCで落ちるが保険）
    t = t.replace("\u3000", " ")
    return t

def _try_parse_datetime(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    cand = s.strip().replace("年", "/").replace("月", "/").replace("日", "")
    cand = cand.replace("-", "/").replace("　", " ")
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(cand, fmt)
            return dt.replace(tzinfo=JST)
        except Exception:
            pass
    return None

def _split_dt_components(dt: Optional[datetime]) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[str], Optional[int], Optional[int]]:
    if not dt:
        return None, None, None, None, None, None
    dt = dt.astimezone(JST)
    return dt.year, dt.month, dt.day, WEEKDAYS_JA[dt.weekday()], dt.hour, dt.minute

def _first_date_yyyymmdd(*vals) -> str:
    for v in vals:
        dt = _try_parse_datetime(v)
        if dt:
            return dt.strftime("%Y%m%d")
    return datetime.now(JST).strftime("%Y%m%d")

def minutes_between(a: Optional[str], b: Optional[str]) -> Optional[int]:
    s = _try_parse_datetime(a); e = _try_parse_datetime(b)
    if s and e:
        return int((e - s).total_seconds() // 60)
    return None

def _split_lines(text: Optional[str], max_lines: int = 5) -> List[str]:
    if not text:
        return []
    lines = [ln.strip() for ln in text.splitlines() if ln.strip() != ""]
    if len(lines) <= max_lines:
        return lines
    return lines[: max_lines - 1] + [lines[max_lines - 1] + "…"]

# ====== 行パーサ版 抽出ロジック（巻き込み防止・堅牢） ======
# ラベル → 保存先キーへの正規化マップ
LABEL_CANON = {
    "管理番号": "管理番号",
    "物件名": "物件名",
    "住所": "住所",
    "窓口会社": "窓口会社",
    "窓口": "窓口会社",  # 同義
    "メーカー": "メーカー",
    "制御方式": "制御方式",
    "契約種別": "契約種別",
    "受信時刻": "受信時刻",
    "通報者": "通報者",
    "現着時刻": "現着時刻",
    "完了時刻": "完了時刻",
    "受信内容": "受信内容",
    "現着状況": "現着状況",
    "原因": "原因",
    "処置内容": "処置内容",
    "対応者": "対応者",
    "完了連絡先1": "完了連絡先1",
    "送信者": "送信者",
    "詳細はこちら": "受付URL",          # URL格納
    "現着・完了登録はこちら": "現着完了登録URL",
    "受付番号": "受付番号",
}
MULTILINE_KEYS = {"受信内容", "現着状況", "原因", "処置内容"}

LABEL_REGEX = re.compile(r"^\s*([^\s:：]+(?:・[^\s:：]+)?)\s*[:：]\s*(.*)$")

def _strip_url_tail(u: str) -> str:
    return re.sub(r"[)\]＞＞）」】>]+$", "", u.strip())

def extract_fields(raw_text: str) -> Dict[str, Optional[str]]:
    t = normalize_text(raw_text)
    lines = t.split("\n")

    # 出力初期化
    out_keys = {
        "管理番号","物件名","住所","窓口会社","メーカー","制御方式","契約種別",
        "受信時刻","通報者","現着時刻","完了時刻",
        "受信内容","現着状況","原因","処置内容",
        "対応者","送信者","受付番号","受付URL","現着完了登録URL",
        "作業時間_分","案件種別(件名)"
    }
    out: Dict[str, Optional[str]] = {k: None for k in out_keys}

    # 件名（任意）
    m_case = re.search(r"^件名:\s*【\s*([^】]+)\s*】", t, flags=re.MULTILINE)
    if m_case:
        out["案件種別(件名)"] = m_case.group(1).strip()
    m_mane = re.search(r"件名:.*?【[^】]+】\s*([A-Z0-9\-]+)", t, flags=re.IGNORECASE)
    subject_manageno = m_mane.group(1).strip() if m_mane else None

    current_multikey: Optional[str] = None
    buffer: List[str] = []
    awaiting_url_for: Optional[str] = None  # "受付URL" or "現着完了登録URL"

    def _flush_buffer():
        nonlocal buffer, current_multikey
        if current_multikey and buffer:
            val = "\n".join([ln for ln in buffer if ln.strip() != ""]).strip()
            out[current_multikey] = val or None
        buffer = []
        current_multikey = None

    i = 0
    while i < len(lines):
        line = lines[i]

        # URL待ち行（ラベル行の次に来るURL）
        if awaiting_url_for and line.strip().startswith("http"):
            out[awaiting_url_for] = _strip_url_tail(line)
            awaiting_url_for = None
            i += 1
            continue

        m = LABEL_REGEX.match(line)
        if m:
            # 直前の複数行ブロックを閉じる
            _flush_buffer()

            raw_label = m.group(1).strip()
            value_part = m.group(2).strip()
            canon = LABEL_CANON.get(raw_label)

            if canon is None:
                # 知らないラベルはスキップ（巻き込み防止）
                i += 1
                continue

            if canon in MULTILINE_KEYS:
                # 複数行ブロック開始：その行の右辺を最初の行として入れる
                current_multikey = canon
                buffer = []
                if value_part:
                    buffer.append(value_part)
            elif canon in ("受付URL", "現着完了登録URL"):
                # URLは同一行か次行に来る
                url = None
                if "http" in value_part:
                    murl = re.search(r"(https?://\S+)", value_part)
                    if murl:
                        url = _strip_url_tail(murl.group(1))
                if url:
                    out[canon] = url
                else:
                    awaiting_url_for = canon  # 次のURL行を待つ
            else:
                # 単行
                if canon == "管理番号" and not value_part and subject_manageno:
                    out[canon] = subject_manageno
                else:
                    out[canon] = value_part or out.get(canon)

            # 受付番号は「詳細はこちら」の行に混在することがあるので、行全体からも拾う
            if "受付番号" in raw_label or "受付番号" in line:
                mnum = re.search(r"受付番号\s*[:：]\s*([0-9]+)", line)
                if mnum:
                    out["受付番号"] = mnum.group(1).strip()

            i += 1
            continue

        # ラベル行ではない
        if current_multikey:
            buffer.append(line)
        else:
            # ラベル外の行に「受付番号:xxxx」が紛れていても拾う
            if out.get("受付番号") is None:
                mnum = re.search(r"受付番号\s*[:：]\s*([0-9]+)", line)
                if mnum:
                    out["受付番号"] = mnum.group(1).strip()
            # URL待ち以外の行は無視（巻き込み防止）
        i += 1

    # 終端フラッシュ
    _flush_buffer()

    # 管理番号の件名補完
    if not out.get("管理番号") and subject_manageno:
        out["管理番号"] = subject_manageno

    # 作業時間（分）
    dur = minutes_between(out.get("現着時刻"), out.get("完了時刻"))
    out["作業時間_分"] = str(dur) if dur is not None and dur >= 0 else None

    return out

# ====== テンプレ書き込み ======
def fill_template_xlsx(template_bytes: bytes, data: Dict[str, Optional[str]]) -> bytes:
    if not template_bytes:
        raise ValueError("テンプレートのバイト列が空です。")

    try:
        wb = load_workbook(io.BytesIO(template_bytes), keep_vba=True)
    except Exception as e:
        raise RuntimeError(f"テンプレートの読み込みに失敗しました（破損の可能性）: {e}") from e

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    def fill_multiline(col_letter: str, start_row: int, text: Optional[str], max_lines: int = 5):
        for i in range(max_lines):
            ws[f"{col_letter}{start_row + i}"] = ""
        if not text:
            return
        for idx, line in enumerate(_split_lines(text, max_lines=max_lines)[:max_lines]):
            ws[f"{col_letter}{start_row + idx}"] = line

    # ---- 単項目
    if data.get("管理番号"): ws["C12"] = data["管理番号"]
    if data.get("メーカー"): ws["J12"] = data["メーカー"]
    if data.get("制御方式"): ws["M12"] = data["制御方式"]
    if data.get("通報者"): ws["C14"] = data["通報者"]
    if data.get("対応者"): ws["L37"] = data["対応者"]

    # 任意：処理修理後
    pa = (st.session_state.get("processing_after") or data.get("処理修理後") or "").strip()
    if pa:
        ws["C35"] = pa

    # 所属
    if data.get("所属"): ws["C37"] = data["所属"]

    # B5/D5/F5 に現在日付（JST）
    now = datetime.now(JST)
    ws["B5"], ws["D5"], ws["F5"] = now.year, now.month, now.day

    # ---- 日時分解ブロック
    def write_dt_block(base_row: int, src_key: str):
        dt = _try_parse_datetime(data.get(src_key))
        y, m, d, wd, hh, mm = _split_dt_components(dt)
        cellmap = {"Y": f"C{base_row}", "Mo": f"F{base_row}", "D": f"H{base_row}",
                   "W": f"J{base_row}", "H": f"M{base_row}", "Min": f"O{base_row}"}
        if y is not None: ws[cellmap["Y"]] = y
        if m is not None: ws[cellmap["Mo"]] = m
        if d is not None: ws[cellmap["D"]] = d
        if wd is not None: ws[cellmap["W"]] = wd
        if hh is not None: ws[cellmap["H"]] = f"{hh:02d}"
        if mm is not None: ws[cellmap["Min"]] = f"{mm:02d}"

    write_dt_block(13, "受信時刻")
    write_dt_block(19, "現着時刻")
    write_dt_block(36, "完了時刻")

    # ---- 複数行
    fill_multiline("C", 15, data.get("受信内容"), max_lines=4)
    fill_multiline("C", 20, data.get("現着状況"))
    fill_multiline("C", 25, data.get("原因"))
    fill_multiline("C", 30, data.get("処置内容"))

    out = io.BytesIO()
    try:
        wb.save(out)
    except Exception as e:
        raise RuntimeError(f"Excel保存時に失敗しました: {e}") from e

    return out.getvalue()

def _sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", name)

def build_filename(data: Dict[str, Optional[str]]) -> str:
    base_day = _first_date_yyyymmdd(data.get("現着時刻"), data.get("完了時刻"), data.get("受信時刻"))
    manageno = _sanitize_filename((data.get("管理番号") or "UNKNOWN").strip().replace("/", "_"))
    bname = _sanitize_filename((data.get("物件名") or "").strip().replace("/", "_"))
    return (f"緊急出動報告書_{manageno}_{bname}_{base_day}.xlsm" if bname
            else f"緊急出動報告書_{manageno}_{base_day}.xlsm")

# ====== Streamlit UI ======
st.set_page_config(page_title=APP_TITLE, layout="centered")
st.markdown(
    """
    <style>
    header {visibility: hidden;}
    .block-container {padding-top: 0rem;}
    .edit-toolbar { position: sticky; top: 0; z-index: 50; backdrop-filter: blur(6px);
      background: rgba(30,30,30,0.08); padding: .5rem .75rem; border-radius: .5rem; margin-bottom: .5rem; }
    .edit-toolbar .btn-row { display: flex; gap: .5rem; align-items: center; flex-wrap: wrap; }
    .edit-badge { font-size: .85rem; background: #ffd24d; color: #4a3b00; padding: .15rem .5rem; border-radius: .5rem; margin-left: .25rem; }
    .missing { color: #b00020; font-weight: 600; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---- セッション初期化 ----
if "step" not in st.session_state: st.session_state.step = 1
if "authed" not in st.session_state: st.session_state.authed = False
if "extracted" not in st.session_state: st.session_state.extracted = None
if "affiliation" not in st.session_state: st.session_state.affiliation = ""
if "template_xlsx_bytes" not in st.session_state: st.session_state.template_xlsx_bytes = None

PASSCODE = _get_passcode()

# Step1: 認証
if st.session_state.step == 1:
    st.subheader("Step 1. パスコード認証")
    if not PASSCODE:
        st.info("（注意）現在、PASSCODEがSecrets/環境変数に未設定です。開発モード想定で空文字として扱います。")
    pw = st.text_input("パスコードを入力してください", type="password")
    if st.button("次へ", use_container_width=True):
        if pw == PASSCODE:
            st.session_state.authed = True
            st.session_state.step = 2
            st.rerun()
        else:
            st.error("パスコードが違います。")

# Step2: 入力
elif st.session_state.step == 2 and st.session_state.authed:
    st.subheader("Step 2. メール本文の貼り付け / 所属 / テンプレ選択")

    template_path = "template.xlsm"
    tpl_col1, tpl_col2 = st.columns([0.55, 0.45])
    with tpl_col1:
        st.caption("① 既定：template.xlsm を探します")
        if os.path.exists(template_path) and not st.session_state.template_xlsx_bytes:
            try:
                with open(template_path, "rb") as f:
                    st.session_state.template_xlsx_bytes = f.read()
                st.success(f"テンプレートを読み込みました: {template_path}")
            except Exception as e:
                st.error(f"テンプレートの読み込みに失敗: {e}")
        elif st.session_state.template_xlsx_bytes:
            st.success("テンプレートは読み込み済みです。")
        else:
            st.warning("既定テンプレートが見つかりません。②のアップロードをご利用ください。")

    with tpl_col2:
        st.caption("② またはテンプレ.xlsmをアップロード")
        up = st.file_uploader("テンプレート（.xlsm）", type=["xlsm"], accept_multiple_files=False)
        if up is not None:
            st.session_state.template_xlsx_bytes = up.read()
            st.success(f"アップロード済み: {up.name}")

    if not st.session_state.template_xlsx_bytes:
        st.error("テンプレートが未準備です。template.xlsm を配置するか、上でアップロードしてください。")
        st.stop()

    aff = st.text_input("所属", value=st.session_state.affiliation)
    st.session_state.affiliation = aff

    processing_after = st.text_input("処理修理後（任意）", value=st.session_state.get("processing_after", ""))
    st.session_state["processing_after"] = processing_after

    text = st.text_area("故障完了メール（本文）を貼り付け", height=240, placeholder="ここにメール本文を貼り付け...")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("抽出する", use_container_width=True):
            if not text.strip():
                st.warning("本文が空です。")
            else:
                st.session_state.extracted = extract_fields(text)
                st.session_state.extracted["所属"] = st.session_state.affiliation  # 空もそのまま
                st.session_state.step = 3
                st.rerun()
    with c2:
        if st.button("クリア", use_container_width=True):
            st.session_state.extracted = None
            st.session_state.affiliation = ""
            st.session_state.processing_after = ""
            st.rerun()

# Step3: 抽出確認→Excel生成
elif st.session_state.step == 3 and st.session_state.authed:
    st.subheader("Step 3. 抽出結果の確認・編集 → Excel生成")

    if "processing_after" in st.session_state and st.session_state.extracted is not None:
        if not st.session_state.extracted.get("_processing_after_initialized"):
            st.session_state.extracted["処理修理後"] = st.session_state.get("processing_after", "")
            st.session_state.extracted["_processing_after_initialized"] = True

    if "edit_mode" not in st.session_state: st.session_state.edit_mode = False
    if "edit_buffer" not in st.session_state: st.session_state.edit_buffer = {}

    st.markdown('<div class="edit-toolbar">', unsafe_allow_html=True)
    tb1, tb2, tb3, tb4 = st.columns([0.22, 0.22, 0.22, 0.34])
    with tb1:
        if not st.session_state.edit_mode:
            if st.button("✏️ 一括編集モードに入る", use_container_width=True):
                _enter_edit_mode(); st.rerun()
        else:
            if st.button("✅ すべて保存", type="primary", use_container_width=True):
                _save_edit(); st.success("保存しました"); st.rerun()
    with tb2:
        if st.session_state.edit_mode:
            if st.button("↩️ 変更を破棄", use_container_width=True):
                _cancel_edit(); st.info("変更を破棄しました"); st.rerun()
        else:
            st.write("")
    with tb3:
        working = _get_working_dict()
        miss = [k for k in REQUIRED_KEYS if _is_required_missing(working, k)]
        st.warning("必須未入力: " + "・".join(miss)) if miss else st.info("必須は入力済み")
    with tb4:
        mode = "ON" if st.session_state.edit_mode else "OFF"
        st.markdown(f"**編集モード:** {mode} " + ("" if not st.session_state.edit_mode else '<span class="edit-badge">一括編集中（指定項目のみ編集可）</span>'),
                    unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    data = _get_working_dict()

    with st.expander("① 編集対象（まとめて編集・すべて必須）", expanded=True):
        render_field("通報者", "通報者", 1, editable_in_bulk=True)
        render_field("受信内容", "受信内容", 4, editable_in_bulk=True)
        render_field("現着状況", "現着状況", 5, editable_in_bulk=True)
        render_field("原因", "原因", 5, editable_in_bulk=True)
        render_field("処置内容", "処置内容", 5, editable_in_bulk=True)
        render_field("処理修理後（Step2入力値）", "処理修理後", 1, editable_in_bulk=True)
        render_field("所属（Step2入力値）", "所属", 1, editable_in_bulk=True)

    with st.expander("② 基本情報（表示）", expanded=True):
        render_field("管理番号", "管理番号", 1)
        render_field("物件名", "物件名", 1)
        render_field("住所", "住所", 2)
        render_field("窓口会社", "窓口会社", 1)
        render_field("制御方式", "制御方式", 1)
        render_field("契約種別", "契約種別", 1)
        render_field("メーカー", "メーカー", 1)

    with st.expander("③ 受付・現着・完了（表示）", expanded=True):
        render_field("受信時刻", "受信時刻", 1)
        render_field("現着時刻", "現着時刻", 1)
        render_field("完了時刻", "完了時刻", 1)

        t_recv_to_arrive = minutes_between(data.get("受信時刻"), data.get("現着時刻"))
        t_work = minutes_between(data.get("現着時刻"), data.get("完了時刻"))
        t_recv_to_done = minutes_between(data.get("受信時刻"), data.get("完了時刻"))

        def _fmt_minutes(v: Optional[int]) -> str:
            return f"{v} 分" if (v is not None and v >= 0) else "—"

        c1, c2, c3 = st.columns(3)
        with c1: st.info(f"受付〜現着時間: {_fmt_minutes(t_recv_to_arrive)}")
        with c2: st.info(f"作業時間: {_fmt_minutes(t_work)}")
        with c3: st.info(f"受付〜完了時間: {_fmt_minutes(t_recv_to_done)}")

    with st.expander("④ その他情報（表示）", expanded=False):
        render_field("対応者", "対応者", 1)
        render_field("送信者", "送信者", 1)
        render_field("受付番号", "受付番号", 1)
        render_field("受付URL", "受付URL", 1)
        render_field("現着完了登録URL", "現着完了登録URL", 1)

    st.divider()

    try:
        is_editing = st.session_state.get("edit_mode", False)
        gen_data = _get_working_dict()
        missing_now = [k for k in REQUIRED_KEYS if _is_required_missing(gen_data, k)]
        can_generate = (not is_editing) and (not missing_now)

        if can_generate:
            xlsx_bytes = fill_template_xlsx(st.session_state.template_xlsx_bytes, gen_data)
            fname = build_filename(gen_data)
            st.download_button(
                "Excelを生成（.xlsm）",
                data=xlsx_bytes,
                file_name=fname,
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                use_container_width=True,
                disabled=False,
                help="一括編集モードはオフ、かつ必須項目がすべて入力されている場合に生成できます",
            )
        else:
            st.download_button(
                "Excelを生成（.xlsm）",
                data=b"",
                file_name="未生成.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                use_container_width=True,
                disabled=True,
                help="一括編集モード中は保存後に生成できます。必須未入力がある場合も生成できません。",
            )
            if is_editing:
                st.warning("一括編集中は生成できません。「✅ すべて保存」を押して編集を確定してください。")
            if missing_now:
                st.error("未入力の必須項目があります： " + "・".join(missing_now))

    except Exception as e:
        st.error(f"テンプレート書き込み中にエラーが発生しました: {e}")
        with st.expander("詳細（開発者向け）"):
            st.code("".join(traceback.format_exception(*sys.exc_info())), language="python")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Step2に戻る", use_container_width=True):
            st.session_state.step = 2; st.rerun()
    with c2:
        if st.button("最初に戻る", use_container_width=True):
            st.session_state.step = 1
            st.session_state.extracted = None
            st.session_state.affiliation = ""
            st.session_state.processing_after = ""
            st.session_state.edit_mode = False
            st.session_state.edit_buffer = {}
            st.rerun()

# 認証未完了時フォールバック
else:
    st.warning("認証が必要です。Step1に戻ります。")
    st.session_state.step = 1
